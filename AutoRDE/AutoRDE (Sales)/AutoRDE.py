from openpyxl import load_workbook
from pynput.keyboard import Key, Controller
from PIL import Image
import pyperclip
import pyautogui
import time
book = load_workbook("Input.xlsx")
sheet = book.active
rows = sheet.rows
keyboard = Controller()

rows = sheet.max_row

print(sheet.max_row)

pyautogui.alert(text='Run the program?', title='AutoRDE', button='Run')
time.sleep(0.5)


for i in range(1, sheet.max_row):

    word = sheet.cell(row=i,column=1).value
    print(word)
    number = sheet.cell(row=i,column=2).value
    print(sheet.cell(row=i,column=2).value)

    while pyautogui.locateCenterOnScreen("RegisteredName.png", confidence = 0.95) == None:
        print("REGoffscreen")

    namecord = pyautogui.locateCenterOnScreen("RegisteredName.png", confidence = 0.95)
    pyautogui.click(namecord)
    pyautogui.typewrite(word)
    pyautogui.press("enter")
    print("works")


    while pyautogui.locateCenterOnScreen("Grosstax.png", confidence = 0.95) == None:
        print("GROSoffscreen")

    numbercord = pyautogui.locateCenterOnScreen("Grosstax.png", confidence = 0.95)
    pyautogui.click(numbercord)
    pyautogui.typewrite(str(number))
    print("works")
 

    while pyautogui.locateCenterOnScreen("Vatrate.png", confidence = 0.9) == None:
        print("VAToffscreen")

    vatcord = pyautogui.locateCenterOnScreen("Vatrate.png", confidence = 0.9)
    pyautogui.click(vatcord)
    pyautogui.write("12")
    print("works")

    while pyautogui.locateCenterOnScreen("save.png", confidence = 0.95) == None:
        print("SAVoffscreen")

    savecord = pyautogui.locateCenterOnScreen("save.png", confidence = 0.95)
    pyautogui.click(savecord)
    print("works")

    while pyautogui.locateCenterOnScreen("add.png", confidence = 0.95) == None:
        print("ADDoffscreen")

    addcord = pyautogui.locateCenterOnScreen("add.png", confidence = 0.95)
    pyautogui.click(addcord)
    print("works")
