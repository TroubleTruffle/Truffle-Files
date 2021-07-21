from openpyxl import load_workbook
from PIL import Image
import pyautogui
import time
from playsound import playsound

book = load_workbook("Input.xlsx")
sheet = book.active
rows = sheet.rows

pyautogui.alert(text='Run the program?', title='AutoRDE', button='Run')
time.sleep(0.5)
print("5")
time.sleep(0.5)
print("4")
time.sleep(0.5)
print("3")
time.sleep(0.5)
print("2")
time.sleep(0.5)
print("1")
time.sleep(0.5)

start = 95
end = 117
for i in range(start, end + 1):
    print("\n\n")
    print("----------------[",start," / ",end ,"]-------------------")

    start += 1

    word = sheet.cell(row=i,column=1).value
    print("NAME:      ", word)

    tin = sheet.cell(row=i,column=2).value
    vtin = str(tin)[0:9]
    print("TIN:       ",vtin[0:3],'-',vtin[3:6],'-',vtin[6:9])

    number = sheet.cell(row=i,column=3).value
    print("AMOUNT:    ",str(number))

    addr = sheet.cell(row=i,column=4).value
    div = len(str(addr)) / 2

    addr1 = str(addr[0:int(div)].lower())
    addr2 = str(addr[int(div):].lower()).strip()

    if len(addr1) > 25:
        while len(addr1) > 25:
            addr1 = addr1[:-2]

    if len(addr2) > 25:
        while len(addr2) > 25:
            addr2 = addr2[:-2]

    print("SUBSTREET: ", addr1)
    print("DISTRICT:  ", addr2)
    print("Errors: \n")

    while pyautogui.locateCenterOnScreen("tin.png", confidence = 0.9) == None:
        print("tinGoffscreen")
        playsound('error.mp3')

    tincord = pyautogui.locateCenterOnScreen("tin.png", confidence = 0.8)
    pyautogui.click(tincord)
    pyautogui.typewrite(str(vtin))

    while pyautogui.locateCenterOnScreen("no.png", confidence = 0.9) != None:
        nocord = pyautogui.locateCenterOnScreen("no.png", confidence = 0.9)
        pyautogui.click(nocord)
        playsound('error.mp3')


    if pyautogui.locateCenterOnScreen("RegisteredName.png", confidence = 0.90) != None:
        namecord = pyautogui.locateCenterOnScreen("RegisteredName.png", confidence = 0.90)
        pyautogui.click(namecord)
        pyautogui.typewrite(word.lower())
        
        if pyautogui.locateCenterOnScreen("Substreet.png", confidence = 0.91) == None:
            print("suboffscreen")
            playsound('error.mp3')

        else:
            subcord = pyautogui.locateCenterOnScreen("Substreet.png", confidence = 0.91)
            pyautogui.click(subcord)
            pyautogui.typewrite(addr1)

        while pyautogui.locateCenterOnScreen("District.png", confidence = 0.90) == None:
            print("disoffscreen")
            pyautogui.press('del')
            time.sleep(0.2)
            playsound('error.mp3')


        discord = pyautogui.locateCenterOnScreen("District.png", confidence = 0.90)
        pyautogui.click(discord)
        pyautogui.typewrite(addr2)

    while pyautogui.locateCenterOnScreen("Grosstax.png", confidence = 0.9) == None:
        print("GROSoffscreen")
        playsound('error.mp3')


    numbercord = pyautogui.locateCenterOnScreen("Grosstax.png", confidence = 0.9)
    pyautogui.click(numbercord)
    pyautogui.typewrite(str(number))

    time.sleep(0.2)
    pyautogui.press(".")
    time.sleep(0.2)
    pyautogui.press("enter")
    time.sleep(0.2)
    pyautogui.press(".")

    if number % 1 == 0:
        pyautogui.press("enter")
        pyautogui.press(".")

    while pyautogui.locateCenterOnScreen("Vatrate.png", confidence = 0.87) == None:
        print("VAToffscreen")
        playsound('error.mp3')


    vatcord = pyautogui.locateCenterOnScreen("Vatrate.png", confidence = 0.8)
    pyautogui.click(vatcord)
    pyautogui.write("12")

    while pyautogui.locateCenterOnScreen("save.png", confidence = 0.95) == None:
        print("SAVoffscreen")
        playsound('error.mp3')


    savecord = pyautogui.locateCenterOnScreen("save.png", confidence = 0.95)
    pyautogui.click(savecord)

    while pyautogui.locateCenterOnScreen("add.png", confidence = 0.95) == None:
        print("ADDoffscreen")
        playsound('error.mp3')
    addcord = pyautogui.locateCenterOnScreen("add.png", confidence = 0.95)
    pyautogui.click(addcord)
    time.sleep(0.2)

playsound('End.mp3')

